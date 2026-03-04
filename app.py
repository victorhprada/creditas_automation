import streamlit as st
import openpyxl
from io import BytesIO
import gc
from openpyxl.formula.translate import Translator
from copy import copy
import datetime

# ---------------------------------------------------------
# SEÇÃO 1: FUNÇÕES DE APOIO E REGRAS DE NEGÓCIO
# ---------------------------------------------------------

def encontrar_ultima_linha(ws, coluna_referencia=1):
    """
    Encontra a verdadeira última linha preenchida baseada na coluna A.
    Previne colar dados no final da planilha por causa de células formatadas vazias.
    """
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=coluna_referencia).value is not None:
            return row
    return 1 # Retorna 1 se estiver totalmente vazia (apenas header)

def copiar_originacao_para_base(ws_parceiro, ws_base):
    """
    Copia os dados da coluna A (1) até Q (17) da aba do Parceiro
    para a primeira linha vazia da aba Base.
    """
    linha_destino_inicio = encontrar_ultima_linha(ws_base) + 1
    linha_destino = linha_destino_inicio
    registros_copiados = 0
    
    # iter_rows com values_only=True é crucial para performance com arquivos de 10MB+
    # min_row=2 pula o cabeçalho do parceiro. max_col=16 pega de A até P.
    for row_values in ws_parceiro.iter_rows(min_row=2, max_col=17):
        
        # Validação de segurança: ignora linhas onde todas as células estão vazias
        # (usamos uma compreensão de lista para checar os valores)
        if not any(cell.value is not None and str(cell.value).strip() != "" for cell in row_values):
            continue
            
        for col_idx, cell_origem in enumerate(row_values, start=1):
            cell_destino = ws_base.cell(row=linha_destino, column=col_idx)
            cell_destino.value = cell_origem.value

            # Copia o formato da célula (Isso é o que salva o CNPJ da notação científica,
            # além de preservar formatação de datas e moedas)
            if cell_origem.has_style:
                cell_destino.number_format = cell_origem.number_format
            
        linha_destino += 1
        registros_copiados += 1

    linha_destino_fim = linha_destino - 1
        
    return linha_destino_inicio, linha_destino_fim, registros_copiados

def preencher_formulas_colunas_r_v(ws_base, linha_inicio, linha_fim):
    """
    Arrasta as fórmulas e valores das colunas R(18) a V(22) da última linha 
    preenchida para as novas linhas inseridas, atualizando as referências.
    """
    linha_referencia = linha_inicio - 1
    
    if linha_referencia < 2:
        raise ValueError("A aba BASE precisa ter pelo menos uma linha de dados com fórmulas para servir de molde.")
        
    colunas_alvo = range(18, 23) # 18=R, 19=S, 20=T, 21=U, 22=V
    
    for row in range(linha_inicio, linha_fim + 1):
        for col in colunas_alvo:
            celula_origem = ws_base.cell(row=linha_referencia, column=col)
            celula_destino = ws_base.cell(row=row, column=col)
            
            # Copia o estilo completo (cor de fundo amarela, bordas, fontes)
            if celula_origem.has_style:
                celula_destino._style = copy(celula_origem._style)
            
            # Se a célula for uma FÓRMULA
            if celula_origem.data_type == 'f':
                nova_formula = Translator(celula_origem.value, origin=celula_origem.coordinate).translate_formula(celula_destino.coordinate)
                celula_destino.value = nova_formula
            # Se for TEXTO ESTÁTICO (ex: "Não pago" ou "setembro")
            else:
                celula_destino.value = celula_origem.value

def copiar_historico_filtrado(ws_origem, ws_destino, mes_filtro, mes_faturamento, ws_nova_aba):
    """
    Filtra a origem e copia os dados para DOIS lugares simultaneamente:
    1. Para a aba 'Parcelas pagas' (com fórmulas extras)
    2. Para a nova aba mensal (apenas colunas A a Q)
    """
    linha_destino = encontrar_ultima_linha(ws_destino) + 1
    linha_destino_nova_aba = encontrar_ultima_linha(ws_nova_aba) + 1
    registros_copiados = 0

    meses_extenso = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho', 
                     7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'}

    # Transforma o input "01-2026" no formato que o Excel guardou: "01/01/2026"
    try:
        mes_input, ano_input = mes_filtro.strip().split('-')
        data_alvo_str = f"01/{mes_input}/{ano_input}"
    except ValueError:
        # Fallback de segurança caso o usuário digite fora do padrão
        data_alvo_str = mes_filtro.strip()
    
    for row in ws_origem.iter_rows(min_row=2, max_col=17):
        
        # Ignora linhas totalmente vazias
        if not any(cell.value is not None and str(cell.value).strip() != "" for cell in row):
            continue
            
        # A coluna Q é a 17ª coluna. No Python (que começa a contar do 0), é o índice 16 da tupla 'row'.
        celula_mes = row[16]
        valor_bruto = celula_mes.value
        valor_mes_celula = ""

        # Padroniza o que vem do Excel para o formato "DD/MM/YYYY"
        if isinstance(valor_bruto, datetime.date):
            valor_mes_celula = valor_bruto.strftime("%d/%m/%Y")
        elif valor_bruto is not None:
            valor_mes_celula = str(valor_bruto).strip().split(" ")[0]
        
        # A CATRACA: Se o mês for igual ao digitado, nós copiamos a linha
        if valor_mes_celula == data_alvo_str:
            for col_idx, cell_origem in enumerate(row, start=1):
                celula_destino = ws_destino.cell(row=linha_destino, column=col_idx)
                celula_nova = ws_nova_aba.cell(row=linha_destino_nova_aba, column=col_idx)
                
                if col_idx == 16 and cell_origem.value is not None:
                    valor_p = cell_origem.value
                    valor_calculado = valor_p

                    if isinstance(valor_p, datetime.date):
                        valor_calculado = meses_extenso.get(valor_p.month, valor_p)
                    elif isinstance(valor_p, str) and "/" in valor_p:
                        try:
                            mes_num = int(valor_p.split("/")[1])
                            valor_calculado = meses_extenso.get(mes_num, valor_p)
                        except:
                            pass
                    celula_destino.value = valor_calculado
                    celula_nova.value = valor_calculado
                else:
                    celula_destino.value = cell_origem.value
                    celula_nova.value = cell_origem.value

                if cell_origem.has_style:
                    celula_destino.number_format = cell_origem.number_format
                    celula_nova.number_format = cell_origem.number_format

            ws_destino.cell(row=linha_destino, column=18).value = f"=N{linha_destino}/M{linha_destino}"
            ws_destino.cell(row=linha_destino, column=18).number_format = "0.00%"
            ws_destino.cell(row=linha_destino, column=19).value = mes_faturamento
            
            linha_destino += 1
            linha_destino_nova_aba += 1
            registros_copiados += 1
            
    return registros_copiados

def gerar_nome_aba_mes(mes_filtro):
    """
    Converte o input '01-2026' no formato de aba 'Jan.26'.
    """
    meses_abrev = {"01": "Jan", "02": "Fev", "03": "Mar", "04": "Abr", "05": "Mai", "06": "Jun",
                   "07": "Jul", "08": "Ago", "09": "Set", "10": "Out", "11": "Nov", "12": "Dez"}
    try:
        mes, ano = mes_filtro.strip().split('-')
        return f"{meses_abrev.get(mes, mes)}.{ano[-2:]}"
    except ValueError:
        return mes_filtro # Fallback caso o usuário digite algo maluco

# ---------------------------------------------------------
# SEÇÃO 2: INTERFACE STREAMLIT
# ---------------------------------------------------------

st.set_page_config(page_title="Automação Creditas", page_icon="📊", layout="centered")
st.title("📊 Processador de Benefícios e Comissionamento")

with st.form("form_processamento"):
    arquivo_parceiro = st.file_uploader("1️⃣ Arquivo do Parceiro (Benefits_Comissionamento_Sênior.xlsx)", type=["xlsx"])
    arquivo_base = st.file_uploader("2️⃣ Arquivo BASE (Acompanhamento creditas base.xlsx)", type=["xlsx", "xlsm"])
    mes_referencia = st.text_input("📅 Mês de Referência (Filtro da Coluna Q)", value="01-2026", help="Digite no formato MM-AAAA")
    mes_faturamento = st.text_input("🏷️ Mês de Faturamento (Coluna S)", value="Janeiro", help="Ex: Janeiro, Fevereiro, etc.")

    submit = st.form_submit_button("Iniciar", type="primary")

if submit:
    if not arquivo_parceiro or not arquivo_base:
        st.error("⚠️ Por favor, envie as duas planilhas antes de processar.")
    else:
        with st.status("🚀 Processando...", expanded=True) as status:
            try:
                st.write("📥 Lendo arquivos na memória (isso pode levar alguns segundos)...")
                # Parceiro carrega apenas valores (data_only=True) para ignorar fórmulas pesadas
                parceiro_wb = openpyxl.load_workbook(arquivo_parceiro, data_only=True)
                # Base carrega com data_only=False para preservar as fórmulas existentes lá dentro
                base_wb = openpyxl.load_workbook(arquivo_base, data_only=False)

                aba_parceiro_nome = "Apoio | Originação e Repasse"
                aba_base_nome = "CREDITAS BASE"

                st.write("⚙️ Validando abas...")
                if aba_parceiro_nome not in parceiro_wb.sheetnames:
                    raise ValueError(f"Aba '{aba_parceiro_nome}' não encontrada no arquivo do PARCEIRO.")
                if aba_base_nome not in base_wb.sheetnames:
                    raise ValueError(f"Aba '{aba_base_nome}' não encontrada no arquivo BASE.")

                ws_parceiro = parceiro_wb[aba_parceiro_nome]
                ws_base = base_wb[aba_base_nome]

                st.write(f"🔄 Copiando dados de '{aba_parceiro_nome}' para '{aba_base_nome}'...")
                linha_inicio, linha_fim, qtd_copiada = copiar_originacao_para_base(ws_parceiro, ws_base)
                st.write(f"✅ {qtd_copiada} linhas copiadas com sucesso!")

                if qtd_copiada > 0:
                    st.write("🔄 Aplicando fórmulas e estilos nas colunas R até V...")
                    preencher_formulas_colunas_r_v(ws_base, linha_inicio, linha_fim)
                    st.write("✅ Fórmulas aplicadas com sucesso!")

                aba_parceiro_hist = "Histórico de relatórios de comi"
                aba_base_parcelas = "Parcelas pagas"

                st.write(f"⚙️ Validando abas da Etapa 2...")
                if aba_parceiro_hist not in parceiro_wb.sheetnames:
                    raise ValueError(f"Aba '{aba_parceiro_hist}' não encontrada no arquivo do PARCEIRO.")
                if aba_base_parcelas not in base_wb.sheetnames:
                    raise ValueError(f"Aba '{aba_base_parcelas}' não encontrada no arquivo BASE.")

                ws_hist = parceiro_wb[aba_parceiro_hist]
                ws_parcelas = base_wb[aba_base_parcelas]

                nome_nova_aba = gerar_nome_aba_mes(mes_referencia)
                st.write(f"🔄 Criando nova aba '{nome_nova_aba}'...")

                if nome_nova_aba not in base_wb.sheetnames:
                    ws_nova = base_wb.create_sheet(nome_nova_aba)
                    for col in range(1, 18):
                        c_origem = ws_hist.cell(row=1, column=col)
                        c_destino = ws_nova.cell(row=1, column=col)
                        c_destino.value = c_origem.value
                        if c_origem.has_style:
                            c_destino._style = copy(c_origem._style)
                        else:
                            ws_nova = base_wb[nome_nova_aba]

                st.write(f"🔄 Filtrando ({mes_referencia}) e copiando para '{aba_base_parcelas}'...")
                qtd_hist = copiar_historico_filtrado(ws_hist, ws_parcelas, mes_referencia, mes_faturamento, ws_nova)
                st.write(f"✅ {qtd_hist} linhas históricas copiadas com sucesso!")

                st.write("💾 Gerando arquivo atualizado para download...")
                output = BytesIO()
                base_wb.save(output)
                output.seek(0)

                # Limpeza severa de memória (obrigatório para nosso cenário de 10MB+)
                del parceiro_wb, base_wb, ws_parceiro, ws_base, ws_hist, ws_parcelas, ws_nova
                gc.collect()

                status.update(label="✅ Processamento Concluído!", state="complete", expanded=False)

                st.download_button(
                    label="📥 Baixar BASE Atualizada",
                    data=output,
                    file_name="CREDITAS_BASE_ATUALIZADA.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            except Exception as e:
                status.update(label="❌ Erro no Processamento", state="error")
                st.error(f"Ocorreu um erro: {str(e)}")